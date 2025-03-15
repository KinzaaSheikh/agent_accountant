import pandas as pd
import fitz
import re
from openpyxl import load_workbook


def main():
    # Load bank statement
    bank_df = pd.read_excel("Input 1 - Bank Statement.xlsx")

    # Load transaction records
    transactions_df = pd.read_excel("Input 2 - Transaction Records.xlsx")

    total_receipts = transactions_df[transactions_df["Amount"] > 0]["Amount"].sum()
    total_payments = transactions_df[transactions_df["Amount"] < 0]["Amount"].sum()
    net_cash_flow = total_receipts + total_payments
    doc = fitz.open("Input 3 - Vouchers.pdf")
    text = ''

    for page in doc:
        text += page.get_text()
    doc.close()

    amount_match = re.search(r"Amount[:\s]+([\d,]+\.?\d*)", text)
    if amount_match:
        amount = float(amount_match.group(1).replace(',', ''))


    # print(f"Total Receipts: {total_receipts}")
    # print(f"Total Payments: {total_payments}")
    # print(f"Net Cash Flow: {net_cash_flow}")
    # print(f"Amount: {amount_match}")

    # def main():
    #     print("Hello from pycash!")

    # Load template workbook (or create new workbook)
    wb = load_workbook("Output - Statement_of_Cash_Receipts_and_Payments_template.xlsx")
    ws = wb.active  # assume data goes in the first sheet

    # Write values to specific cells (assuming template has known cell positions)
    ws["B2"] = total_receipts    # e.g., cell B2 for Total Cash Receipts amount
    ws["B3"] = total_payments    # cell B3 for Total Cash Payments amount
    ws["B4"] = net_cash_flow     # Net increase/decrease
    # ws["B5"] = opening_balance
    # ws["B6"] = closing_balance

    wb.save("Output - Statement_of_Cash_Receipts_and_Payments.xlsx")
    print("Statement saved successfully.")


if __name__ == "__main__":
    main()
